import trgep_toolbox as trgeptb
import numpy as np
from pymoo.visualization.scatter import Scatter
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl

# PRINT RESULT
def print_result(problem,res):
    print('Time: ', res.exec_time)
    if not problem.n_obj == 1:
        print_multi(res)
        draw(problem,res)
        #draw_pop(problem,res)
    else:
        print_one(res)

def print_one(res):
    print("Best solution found ever : \nF = %s " % (res.F))
    x = np.array([res.X[f"x{k:02}"] for k in range(0, 352)])
    print(trgeptb.const_check_debug(x,trgeptb.data))
    
    #df = pd.DataFrame(x)
    #df.to_excel('best.xlsx', index=False, header=False)
    x = trgeptb.save_matrix(x)
    #wb = openpyxl.load_workbook('cost-opt.xlsx')
    wb = openpyxl.load_workbook('em-opt.xlsx')
    sheet = wb['Sheet1']
    #print(sheet.max_row)
    df = pd.DataFrame(x)
    #with pd.ExcelWriter("cost-opt.xlsx", mode="a", engine="openpyxl",if_sheet_exists='overlay') as writer:
    with pd.ExcelWriter("em-opt.xlsx", mode="a", engine="openpyxl",if_sheet_exists='overlay') as writer:
        df.to_excel(writer, index=False, header=False, sheet_name="Sheet1", startrow=sheet.max_row)  
    

def print_multi(res):
    for t in range(len(res.F)):
        print(len(res.F))
        print(str(res.F[t][0])[:3] , "e" , len(str(int(res.F[t][0])))-3 , "          " , str(res.F[t][1])[:3] , "e" , len(str(int(res.F[t][1])))-3)
        x = np.array([res.X[t][f"x{k:02}"] for k in range(0, 352)])
        trgeptb.const_check_debug(x,trgeptb.data)

    val = res.algorithm.callback.data["best"]
    plt.suptitle('Convergence')
    plt.ylabel("best at gen")
    plt.xlabel("gen")
    plt.plot(np.arange(len(val)), val)
    plt.show()

    x = trgeptb.save_matrix(x)
    wb = openpyxl.load_workbook('best.xlsx')
    sheet = wb['Sheet1']
    #print(sheet.max_row)
    df = pd.DataFrame(x)
    with pd.ExcelWriter("best.xlsx", mode="a", engine="openpyxl",if_sheet_exists='overlay') as writer:
        df.to_excel(writer, index=False, header=False, sheet_name="Sheet1", startrow=sheet.max_row)  
    
    
def divide_feasible(res):   
    f_pop_res = []
    if_pop_res = []
    for ind in res.pop:
        nucp,capp,demp,pdemp,climp,capConstraintDict,nuclearInd = trgeptb.const_check(np.array(list(ind.X.values())),trgeptb.data)
        total = nucp+capp+demp+pdemp+climp
        if total == 0:
            f_pop_res.append(ind.F)
        else:
            if_pop_res.append(ind.F)
    return np.array(f_pop_res),np.array(if_pop_res)

def divide_feasible_init(pop):   
    f_pop_res = []
    if_pop_res = []
    for ind in pop:
        nucp,capp,demp,pdemp,climp,capConstraintDict,nuclearInd = trgeptb.const_check(np.array(list(ind.X.values())),trgeptb.data)
        total = nucp+capp+demp+pdemp+climp
        if total == 0:
            f_pop_res.append(ind.F)
        else:
            if_pop_res.append(ind.F)
    return np.array(f_pop_res),np.array(if_pop_res)

def draw(problem,res):
    f_pop_res,if_pop_res = divide_feasible(res)
    plot = Scatter()
    #plot.add(problem.pareto_front(), plot_type="line", color="black", alpha=0.7)
    #plot.add(res.F, plot_type="plot", color="purple", alpha=0.9)
    plot.add(res.F, facecolor="cyan", edgecolor="cyan")
    if len(f_pop_res) > 0:
        plot.add(f_pop_res, facecolor="none", edgecolor="blue")
    if len(if_pop_res) > 0:
        plot.add(if_pop_res, facecolor="none", edgecolor="red")
    plot.add(trgeptb.pareto_list(),facecolor="none", edgecolor="green")
    plot.show() 
    t = res.algorithm
    path = r'picoutputs'
    plot.save(f"{path}/nsga2-POP_{problem.NPOP}-OFFS_{res.algorithm.n_offsprings}-GEN_{res.algorithm.n_gen-1}-SEED_{res.algorithm.seed}-EXEC_{res.exec_time}.png")
    print("")

def draw_pop(problem,pop):
    f_pop_res,if_pop_res = divide_feasible_init(pop)
    plot = Scatter()
    plot.add(problem.pareto_front(), plot_type="line", color="black", alpha=0.7)
    if len(f_pop_res) > 0:
        plot.add(f_pop_res, facecolor="none", edgecolor="blue")
    if len(if_pop_res) > 0:
        plot.add(if_pop_res, facecolor="none", edgecolor="red")
    plot.show()